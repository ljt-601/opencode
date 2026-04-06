#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


DETAIL_PATTERN = re.compile(
  r"^(?P<owner>[^+＋]+)[+＋](?P<dept>[^+＋]+)[+＋](?P<amount>\d+(?:\.\d+)?)(?:[（(](?P<remark>.*?)[）)])?$"
)
COST_PATTERN = re.compile(r"^(?:硬?成本)[:：]?\s*(?P<cost>\d+(?:\.\d+)?)$")
@dataclass
class Record:
  customer_name: str
  owner_name: str
  dept_name: str
  amount_value: Decimal
  cost_value: Optional[Decimal]
  profit_value: Decimal
  biz_type: str
  record_period: Optional[str]
  remark_text: Optional[str]
  contract_text: Optional[str]
  date_text: str
def get_args() -> argparse.Namespace:
  parser = argparse.ArgumentParser(description="将粘贴文本追加写入 Excel")
  parser.add_argument("--excel", required=True, help="目标 Excel 文件路径")
  parser.add_argument("--text-file", required=True, help="待解析文本文件路径")
  parser.add_argument("--sheet", default="表格", help="工作表名称，默认：表格")
  return parser.parse_args()
def parse_decimal(raw_text: str) -> Decimal:
  try:
    return Decimal(raw_text)
  except InvalidOperation as exc:
    raise ValueError(f"无法识别数字：{raw_text}") from exc
def normalize_lines(raw_text: str) -> list[str]:
  lines = [line.strip() for line in raw_text.splitlines()]
  return [line for line in lines if line]
def split_blocks(lines: list[str]) -> list[list[str]]:
  blocks: list[list[str]] = []
  current_block: list[str] = []

  for line in lines:
    if "+" in line or "＋" in line or COST_PATTERN.match(line):
      current_block.append(line)
      continue

    if current_block:
      blocks.append(current_block)
    current_block = [line]

  if current_block:
    blocks.append(current_block)

  return blocks


def infer_biz_type(remark_text: Optional[str], detail_line: str) -> str:
  source_text = f"{detail_line} {remark_text or ''}"
  if any(keyword in source_text for keyword in ["续费", "续签", "地址续费"]):
    return "续签"
  return "新签"


def parse_block(block: list[str], today_text: str) -> Record:
  if len(block) < 2:
    raise ValueError(f"记录不完整：{' | '.join(block)}")

  customer_name = block[0]
  detail_line = block[1]
  detail_match = DETAIL_PATTERN.match(detail_line)
  if not detail_match:
    raise ValueError(f"无法识别明细行：{detail_line}")

  owner_name = detail_match.group("owner").strip()
  dept_name = detail_match.group("dept").strip()
  amount_value = parse_decimal(detail_match.group("amount"))
  remark_text = detail_match.group("remark").strip() if detail_match.group("remark") else None

  cost_value: Optional[Decimal] = None
  record_period: Optional[str] = None
  contract_text: Optional[str] = None

  for extra_line in block[2:]:
    cost_match = COST_PATTERN.match(extra_line)
    if cost_match:
      cost_value = parse_decimal(cost_match.group("cost"))
      continue
    if not record_period and re.search(r"\d{4}[.-]\d", extra_line):
      record_period = extra_line
      continue
    remark_text = extra_line if not remark_text else f"{remark_text}；{extra_line}"

  profit_value = amount_value - cost_value if cost_value is not None else amount_value
  biz_type = infer_biz_type(remark_text, detail_line)

  return Record(
    customer_name=customer_name,
    owner_name=owner_name,
    dept_name=dept_name,
    amount_value=amount_value,
    cost_value=cost_value,
    profit_value=profit_value,
    biz_type=biz_type,
    record_period=record_period,
    remark_text=remark_text,
    contract_text=contract_text,
    date_text=today_text,
  )


def decimal_to_number(value: Optional[Decimal]):
  if value is None:
    return None
  if value == value.to_integral_value():
    return int(value)
  return float(value)


def copy_row_style(ws, source_row: int, target_row: int, start_col: int = 1, end_col: int = 12) -> None:
  for col in range(start_col, end_col + 1):
    source_cell = ws.cell(source_row, col)
    target_cell = ws.cell(target_row, col)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format
  ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def ensure_header(ws) -> None:
  if not ws["A1"].value:
    ws["A1"] = "A(序号)"


def append_records(excel_path: Path, sheet_name: str, records: list[Record]) -> list[tuple[int, Record]]:
  workbook = load_workbook(excel_path)
  worksheet = workbook[sheet_name]
  ensure_header(worksheet)

  result_rows: list[tuple[int, Record]] = []
  for record in records:
    next_row = worksheet.max_row + 1
    style_row = max(2, next_row - 1)
    if next_row > 2:
      copy_row_style(worksheet, style_row, next_row)

    worksheet.cell(next_row, 1, next_row - 1)
    worksheet.cell(next_row, 2, record.date_text)
    worksheet.cell(next_row, 3, record.dept_name)
    worksheet.cell(next_row, 4, record.owner_name)
    worksheet.cell(next_row, 5, record.customer_name)
    worksheet.cell(next_row, 6, decimal_to_number(record.amount_value))
    worksheet.cell(next_row, 7, decimal_to_number(record.cost_value))
    worksheet.cell(next_row, 8, decimal_to_number(record.profit_value))
    worksheet.cell(next_row, 9, record.biz_type)
    worksheet.cell(next_row, 10, record.record_period)
    worksheet.cell(next_row, 11, record.remark_text)
    worksheet.cell(next_row, 12, record.contract_text)
    result_rows.append((next_row, record))

  worksheet.auto_filter.ref = f"A1:L{worksheet.max_row}"
  workbook.save(excel_path)
  return result_rows


def main() -> None:
  args = get_args()
  excel_path = Path(args.excel).expanduser()
  text_path = Path(args.text_file).expanduser()

  if not excel_path.exists():
    raise FileNotFoundError(f"Excel 文件不存在：{excel_path}")
  if not text_path.exists():
    raise FileNotFoundError(f"文本文件不存在：{text_path}")

  raw_text = text_path.read_text(encoding="utf-8")
  lines = normalize_lines(raw_text)
  if not lines:
    raise ValueError("文本内容为空")

  today_text = date.today().isoformat()
  blocks = split_blocks(lines)
  records = [parse_block(block, today_text) for block in blocks]
  result_rows = append_records(excel_path, args.sheet, records)

  for row_number, record in result_rows:
    print(
      f"row={row_number}\t客户={record.customer_name}\t部门={record.dept_name}\t经手人={record.owner_name}"
      f"\t金额={decimal_to_number(record.amount_value)}\t成本={decimal_to_number(record.cost_value)}"
      f"\t利润={decimal_to_number(record.profit_value)}\t备注={record.remark_text or ''}"
    )


if __name__ == "__main__":
  main()
